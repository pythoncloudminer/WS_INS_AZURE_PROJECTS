import logging
import os
from typing import Dict, List, Tuple

import azure.functions as func
import pandas as pd
from azure.storage.blob import BlobServiceClient
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
app = func.FunctionApp()

@app.blob_trigger(arg_name="myblob", path="source",
                               connection="synapsetestio_STORAGE") 
def scqt_cleanser(myblob: func.InputStream):
    """Azure Blob Trigger function that processes Excel files."""
    logger.info(f"Python blob trigger function processed blob \n"
                f"Name: {myblob.name}\n"
                f"Blob Size: {myblob.length} bytes")

    try:
        if not is_excel_file(myblob.name):
            logger.info(f"Skipping non-Excel file: {myblob.name}")
            return

        process_excel_file(myblob)

    except Exception as e:
        logger.error(f"Error processing file {myblob.name}: {str(e)}", exc_info=True)
        raise


def is_excel_file(filename: str) -> bool:
    """Check if the file is an Excel file (.xls or .xlsx)."""
    return filename.lower().endswith(('.xls', '.xlsx'))


def process_excel_file(blob_stream: func.InputStream):
    """Process an Excel file from blob storage."""
    temp_file_path = f"/tmp/{os.path.basename(blob_stream.name)}"
    with open(temp_file_path, "wb") as f:
        f.write(blob_stream.read())

    try:
        workbook = load_workbook(temp_file_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            logger.info(f"Processing worksheet: {sheet_name}")
            worksheet = workbook[sheet_name]

            # Detect and extract the single table from the worksheet
            table = detect_table_in_worksheet(worksheet)
            if not table:
                logger.info(f"No table found in worksheet: {sheet_name}")
                continue

            df = table_to_dataframe(table)
            df = transform_dataframe(df, sheet_name, blob_stream.name)

            save_dataframe_to_blob(
                df=df,
                original_blob_name=blob_stream.name,
                sheet_name=sheet_name
            )

    finally:
        if 'workbook' in locals():
            workbook.close()

        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)


def detect_table_in_worksheet(worksheet: Worksheet) -> Dict[str, List]:
    """
    Detect a single table in an Excel worksheet based on:
    - Header row: First row where 80% of columns have values
    - End row: First row after header where 80% of columns are empty
    """
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}

    num_cols = len(rows[0]) if rows else 0
    if num_cols == 0:
        return {}

    # Find header row (first row with >=80% non-empty cells)
    header_row = None
    for row_idx, row in enumerate(rows):
        non_empty = sum(1 for cell in row if cell is not None and str(cell).strip() != '')
        if non_empty >= 0.8 * num_cols:
            header_row = row_idx
            break

    if header_row is None:
        return {}

    # Find end row (first row after header with >=80% empty cells)
    end_row = len(rows) - 1
    for row_idx in range(header_row + 1, len(rows)):
        row = rows[row_idx]
        empty = sum(1 for cell in row if cell is None or str(cell).strip() == '')
        if empty >= 0.8 * num_cols:
            end_row = row_idx - 1
            break

    if end_row <= header_row:
        return {}

    # Extract all columns (assuming table uses all columns)
    headers = []
    for col_idx in range(num_cols):
        header = rows[header_row][col_idx]
        headers.append(str(header) if header is not None else f"Column_{col_idx + 1}")

    data = []
    for row_idx in range(header_row + 1, end_row + 1):
        row_data = []
        for col_idx in range(num_cols):
            row_data.append(rows[row_idx][col_idx])
        data.append(row_data)

    return {'headers': headers, 'data': data}


def table_to_dataframe(table: Dict[str, List]) -> pd.DataFrame:
    """Convert a detected table to a pandas DataFrame."""
    return pd.DataFrame(table['data'], columns=table['headers'])


def transform_dataframe(df: pd.DataFrame, sheet_name: str, filename: str) -> pd.DataFrame:
    """Apply minimal required transformations."""
    df.columns = [col.strip() for col in df.columns]
    df['source_file'] = os.path.basename(filename)
    df['source_sheet'] = sheet_name
    return df


def save_dataframe_to_blob(df: pd.DataFrame, original_blob_name: str, sheet_name: str):
    """Save a DataFrame to a CSV file in blob storage."""
    original_name = os.path.splitext(os.path.basename(original_blob_name))[0]
    output_blob_name = f"processed/{original_name}_{sheet_name}.csv"
    
    blob_service_client = BlobServiceClient.from_connection_string(
        '<CONNECTION_STRING>'
    )
    
    csv_data = df.to_csv(index=False).encode('utf-8')
    
    blob_client = blob_service_client.get_blob_client(
        container='output',
        blob=output_blob_name
    )
    
    blob_client.upload_blob(csv_data, overwrite=True)
    logger.info(f"Uploaded CSV to blob: {output_blob_name}")
