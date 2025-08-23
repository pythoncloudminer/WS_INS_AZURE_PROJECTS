import logging
import os
import json
import tempfile
from typing import Dict, List

import azure.functions as func
import pandas as pd
from azure.storage.blob import BlobServiceClient
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import xlrd  # make sure xlrd==1.2.0 is installed

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
app = func.FunctionApp()


@app.event_grid_trigger(arg_name="event")
def scqt_cleanser(event: func.EventGridEvent):
    """Azure Event Grid Trigger function that processes Excel files from blob events."""

    try:
        # Event Grid event -> extract blob URL
        event_data = event.get_json()
        blob_url = event_data["url"]
        logger.info(f"Event received for blob: {blob_url}")

        # Get blob client
        blob_service_client = BlobServiceClient.from_connection_string(
            os.getenv("AzureWebJobsStorage")
        )
        blob_client = blob_service_client.get_blob_client(blob_url)

        blob_name = os.path.basename(blob_url)
        if not is_excel_file(blob_name):
            logger.info(f"Skipping non-Excel file: {blob_name}")
            return

        # Download blob to a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(blob_name)[1]) as tmp_file:
            tmp_file.write(blob_client.download_blob().readall())
            temp_file_path = tmp_file.name

        process_excel_file(temp_file_path, blob_name)

    except Exception as e:
        logger.error(f"Error processing event: {str(e)}", exc_info=True)
        raise


def is_excel_file(filename: str) -> bool:
    """Check if the file is an Excel file."""
    return filename.lower().endswith(('.xls', '.xlsx'))


def process_excel_file(temp_file_path: str, blob_name: str):
    """Open Excel (.xls or .xlsx), detect tables, convert to CSV, and save to blob storage."""
    try:
        if blob_name.lower().endswith(".xlsx"):
            # Use openpyxl for .xlsx
            workbook = load_workbook(temp_file_path, read_only=True, data_only=True)
            for sheet_name in workbook.sheetnames:
                logger.info(f"Processing worksheet (xlsx): {sheet_name}")
                worksheet = workbook[sheet_name]

                table = detect_table_in_worksheet(worksheet)
                if not table:
                    logger.info(f"No table found in worksheet: {sheet_name}")
                    continue

                df = table_to_dataframe(table)
                df = transform_dataframe(df, sheet_name, blob_name)
                save_dataframe_to_blob(df, blob_name, sheet_name)

            workbook.close()

        elif blob_name.lower().endswith(".xls"):
            # Use xlrd for .xls
            book = xlrd.open_workbook(temp_file_path)
            for sheet_name in book.sheet_names():
                logger.info(f"Processing worksheet (xls): {sheet_name}")
                sheet = book.sheet_by_name(sheet_name)

                rows = [sheet.row_values(row_idx) for row_idx in range(sheet.nrows)]
                table = detect_table_from_rows(rows)
                if not table:
                    logger.info(f"No table found in worksheet: {sheet_name}")
                    continue

                df = table_to_dataframe(table)
                df = transform_dataframe(df, sheet_name, blob_name)
                save_dataframe_to_blob(df, blob_name, sheet_name)

    finally:
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)


def detect_table_in_worksheet(worksheet: Worksheet) -> Dict[str, List]:
    """Detect a table inside an Excel worksheet (.xlsx)."""
    rows = list(worksheet.iter_rows(values_only=True))
    return detect_table_from_rows(rows)


def detect_table_from_rows(rows: List[List]) -> Dict[str, List]:
    """Detect a table inside an Excel sheet using raw rows (works for .xls)."""
    if not rows:
        return {}

    num_cols = len(rows[0])
    if num_cols == 0:
        return {}

    # Find header row
    header_row = None
    for row_idx, row in enumerate(rows):
        non_empty = sum(1 for cell in row if str(cell).strip() != "")
        if non_empty >= 0.8 * num_cols:
            header_row = row_idx
            break
    if header_row is None:
        return {}

    # Find end row
    end_row = len(rows) - 1
    for row_idx in range(header_row + 1, len(rows)):
        empty = sum(1 for cell in rows[row_idx] if str(cell).strip() == "")
        if empty >= 0.8 * num_cols:
            end_row = row_idx - 1
            break
    if end_row <= header_row:
        return {}

    # Headers
    headers = [
        str(rows[header_row][col_idx]) if rows[header_row][col_idx] else f"Column_{col_idx + 1}"
        for col_idx in range(num_cols)
    ]

    # Data
    data = [
        [rows[row_idx][col_idx] for col_idx in range(num_cols)]
        for row_idx in range(header_row + 1, end_row + 1)
    ]

    return {"headers": headers, "data": data}


def table_to_dataframe(table: Dict[str, List]) -> pd.DataFrame:
    """Convert detected table to pandas DataFrame."""
    return pd.DataFrame(table['data'], columns=table['headers'])


def transform_dataframe(df: pd.DataFrame, sheet_name: str, filename: str) -> pd.DataFrame:
    """Transform DataFrame with additional metadata."""
    df.columns = [col.strip() for col in df.columns]
    df['source_file'] = filename
    df['source_sheet'] = sheet_name
    return df


def save_dataframe_to_blob(df: pd.DataFrame, original_blob_name: str, sheet_name: str):
    """Save DataFrame to blob storage as CSV."""
    original_name = os.path.splitext(os.path.basename(original_blob_name))[0]
    output_blob_name = f"processed/{original_name}_{sheet_name}.csv"

    blob_service_client = BlobServiceClient.from_connection_string(
        os.getenv("AzureWebJobsStorage")
    )

    csv_data = df.to_csv(index=False).encode('utf-8')
    blob_client = blob_service_client.get_blob_client(
        container="output",
        blob=output_blob_name
    )

    blob_client.upload_blob(csv_data, overwrite=True)
    logger.info(f"Uploaded CSV to blob: {output_blob_name}")
